from __future__ import annotations

import csv
import json
import re
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session


MAX_IMPORT_ROWS = 20000


def _sanitize_identifier(value: str, fallback: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]+", "_", value.strip().lower())
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    if not cleaned:
        return fallback
    if cleaned[0].isdigit():
        cleaned = f"_{cleaned}"
    return cleaned


def _parse_csv(data: bytes) -> list[dict[str, Any]]:
    decoded = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            decoded = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise ValueError("Could not decode CSV file")

    reader = csv.DictReader(StringIO(decoded))
    rows = [dict(row) for row in reader]
    return rows


def _parse_json(data: bytes) -> list[dict[str, Any]]:
    payload = json.loads(data.decode("utf-8"))

    if isinstance(payload, dict) and isinstance(payload.get("data"), list):
        payload = payload["data"]

    if not isinstance(payload, list):
        raise ValueError("JSON file must be an array of objects or {\"data\": [...]} format")

    rows: list[dict[str, Any]] = []
    for item in payload:
        if not isinstance(item, dict):
            raise ValueError("JSON array items must be objects")
        rows.append(item)
    return rows


def _parse_xlsx(data: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    raw_rows = list(sheet.iter_rows(values_only=True))
    if not raw_rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in raw_rows[0]]
    normalized_headers = []
    for index, header in enumerate(headers):
        normalized_headers.append(header or f"column_{index + 1}")

    rows: list[dict[str, Any]] = []
    for raw in raw_rows[1:]:
        row: dict[str, Any] = {}
        for index, key in enumerate(normalized_headers):
            row[key] = raw[index] if index < len(raw) else None
        rows.append(row)
    return rows


def _normalize_rows(rows: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    if not rows:
        return [], []

    original_columns: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in original_columns:
                original_columns.append(key)

    used: dict[str, int] = {}
    columns: list[str] = []
    mapping: dict[str, str] = {}

    for index, original in enumerate(original_columns):
        base = _sanitize_identifier(str(original), f"column_{index + 1}")
        current = base
        if current in used:
            used[current] += 1
            current = f"{current}_{used[base]}"
        else:
            used[current] = 1
        mapping[str(original)] = current
        columns.append(current)

    normalized: list[dict[str, Any]] = []
    for row in rows:
        item: dict[str, Any] = {column: None for column in columns}
        for original_key, value in row.items():
            item[mapping[str(original_key)]] = value
        normalized.append(item)

    return columns, normalized


def _to_int(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str) and re.fullmatch(r"[-+]?\d+", value.strip()):
        return int(value.strip())
    return None


def _to_float(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.strip())
        except ValueError:
            return None
    return None


def _to_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"true", "t", "yes", "y", "1"}:
            return True
        if lowered in {"false", "f", "no", "n", "0"}:
            return False
    return None


def _to_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        text_value = value.strip()
        if not text_value:
            return None
        text_value = text_value.replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(text_value)
        except ValueError:
            return None
    return None


def _infer_column_type(values: list[Any]) -> str:
    non_empty = [value for value in values if value not in (None, "")]
    if not non_empty:
        return "TEXT"

    if all(_to_bool(value) is not None for value in non_empty):
        return "BOOLEAN"
    if all(_to_int(value) is not None for value in non_empty):
        return "BIGINT"
    if all(_to_float(value) is not None for value in non_empty):
        return "DOUBLE PRECISION"
    if all(_to_datetime(value) is not None for value in non_empty):
        return "TIMESTAMP"
    return "TEXT"


def _convert_value(value: Any, data_type: str) -> Any:
    if value in (None, ""):
        return None
    if data_type == "BOOLEAN":
        return _to_bool(value)
    if data_type == "BIGINT":
        return _to_int(value)
    if data_type == "DOUBLE PRECISION":
        return _to_float(value)
    if data_type == "TIMESTAMP":
        return _to_datetime(value)
    return str(value)


def import_uploaded_dataset(
    db: Session,
    *,
    filename: str,
    content: bytes,
    user_id: int,
    requested_table_name: str | None = None,
) -> dict[str, Any]:
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if extension == "csv":
        rows = _parse_csv(content)
    elif extension in {"xlsx", "xlsm"}:
        rows = _parse_xlsx(content)
    elif extension == "json":
        rows = _parse_json(content)
    else:
        raise ValueError("Unsupported file type. Please upload CSV, XLSX, or JSON.")

    if not rows:
        raise ValueError("Uploaded file has no rows.")
    if len(rows) > MAX_IMPORT_ROWS:
        raise ValueError(f"File has too many rows. Limit is {MAX_IMPORT_ROWS}.")

    default_name = _sanitize_identifier(filename.rsplit(".", 1)[0], "dataset")
    provided = _sanitize_identifier(requested_table_name or "", "") if requested_table_name else ""
    base = provided or default_name or "dataset"
    table_name = _sanitize_identifier(f"u{user_id}_{base}", f"u{user_id}_dataset")

    columns, normalized_rows = _normalize_rows(rows)
    if not columns:
        raise ValueError("Could not detect columns in uploaded file.")

    column_types: dict[str, str] = {}
    for column in columns:
        column_types[column] = _infer_column_type([row.get(column) for row in normalized_rows])

    columns_sql = ", ".join([f'"{column}" {column_types[column]}' for column in columns])
    db.execute(text(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({columns_sql})'))
    db.execute(text(f'TRUNCATE TABLE "{table_name}"'))

    insert_sql = (
        f'INSERT INTO "{table_name}" ({", ".join([f"\"{column}\"" for column in columns])}) '
        f'VALUES ({", ".join([f":{column}" for column in columns])})'
    )

    payload: list[dict[str, Any]] = []
    for row in normalized_rows:
        item: dict[str, Any] = {}
        for column in columns:
            item[column] = _convert_value(row.get(column), column_types[column])
        payload.append(item)

    db.execute(text(insert_sql), payload)
    db.commit()

    return {
        "table_name": table_name,
        "row_count": len(payload),
        "columns": [{"name": column, "type": column_types[column]} for column in columns],
    }